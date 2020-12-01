#!/usr/bin/python3
# -*- coding: utf-8 -*-

__author__ = 'Rodrigo Mansilha'
__email__ = 'mansilha@unipampa.edu.br'
__version__ = '{0}.{0}.{1}'
__credits__ = ['Comissão Local de Pesquisa (CLP) do Campus Alegrete da Univ. Federal do Pampa - Unipampa']


# Bibliotecas gerais
import sys, os	# system
import argparse # parse arguments
import glob		# Unix style pathname pattern expansion
import logging  # organizar saídas em formato de log
from pathlib import Path
from io import StringIO

# Outras bibliotecas
from openpyxl import load_workbook, Workbook #ler e escrever arquivos em formato do Excel
import bibtexparser #ler e escrever arquivos em formato bibtex   #pip3 install bibtexparser
from tika import parser as tika_parser # ler texto de pdf

from bibtexparser.bparser import BibTexParser
from bibtexparser.customization import convert_to_unicode

# ARQUIVOS DE SAÍDA
PLANILHA_ARTIGOS_PADRAO = "Artigos.xlsx"
PLANILHA_AUTORES_PADRAO = "Autores.xlsx"
PLANILHA_REFERENCIAS_PADRAO = "Referencias.xlsx"
PLANILHA_SECOES_PADRAO = "Secoes.xlsx" # Não utilizada

# DIRETÓRIO DE ENTRADA
DIRETORIO_ENTRADA_PADRAO = "./papers"

# CONFIGURAÇÕES PADRÃO
LANGUAGE_PADRAO = "pt"
SECAO_ABREV_PADRAO = "ART"
AUTHOR_COUNTRY_PADRAO = "Brasil"
PLANILHA_PADRAO = "Planilha1"

# CAMPOS DAS TABELAS
CAMPOS_ARTIGOS = [
		"seq",
		"language",
		"sectionAbbrev",
		"title",
		"titleEn",
		"abstract",
		"abstractEn",
		"keywords",
		"keywordsEn",
		"pages",
		"fileLabel",
		"fileLink"
]

CAMPOS_AUTORES = [
	"article",
	"authorFirstname",
	"authorMiddlename",
	"authorLastname",
	"authorAffiliation",
	"authorAffiliationEn",
	"authorCountry",
	"authorEmail",
	"orcid",
	"authorBio",
	"authorBioEn"
]

CAMPOS_SECOES = [
	"sectionTitle",
	"sectionTitleEn",
	"sectionAbbrev"
]

CAMPOS_REFERENCIAS = [
	"article",
	"references"
]


class Secao(object):

	def __init__(self, sectionAbbrev_=None):
		self.sectionTitle = ""
		self.sectionTitleEn = ""
		self.sectionAbbrev = ""
		if sectionAbbrev_ is not None:
			self.sectionAbbrev = sectionAbbrev_


class Referencia(object):

	def __init__(self, artigo_seq_, referencia_):
		self.article = artigo_seq_
		self.references = referencia_
		# self.DOI = "" # TODO

	def __str__(self):
		msg = "\n"
		msg += "\t\tReferência\n"
		msg += "\t\t----------\n"
		msg += "\t\t\tarticle: %d\n" % self.article
		msg += "\t\t\treferences:%s\n" % self.references
		msg += '\n'
		return msg


class Autor(object):

	id_contador = 0

	def __init__(self, artigo_seq_, autor_str_ = None, author_country_ = None):
		Autor.id_contador += 1
		self.article = artigo_seq_
		self.authorFirstname = ""
		self.authorMiddlename = ""
		self.authorLastname = ""
		self.authorAffiliation = ""
		self.authorAffiliationEn = ""
		self.authorCountry = ""
		self.authorEmail = ""
		self.orcid = ""
		self.authorBio = ""
		self.authorBioEn = ""

		if not autor_str_ is None:
			self.authorFirstname = autor_str_.split(" ")[0]
			self.authorLastname = " ".join(autor_str_.split(" ")[1:])

		if author_country_ is None:
			authorCountry = AUTHOR_COUNTRY_PADRAO

	def __str__(self):
		msg = "\n"
		msg += "\t\tAutor\n"
		msg += "\t\t-----\n"
		msg += "\t\t\tarticle: %d\n" % self.article
		msg += "\t\t\tauthorFirstname:%s\n" % self.authorFirstname
		msg += "\t\t\tauthorLastname:%s\n" % self.authorLastname
		msg += '\n'
		return msg


class Artigo(object):

	def __init__(self, language_, section_abbrev_, seq_, bib_database_=None, pdf_text_=None, pdf_file_name_=None):
		self.seq = seq_
		self.language = language_
		self.sectionAbbrev = section_abbrev_
		self.title = ""
		self.titleEn = ""
		self.abstract = ""
		self.abstractEn = ""
		self.keywords = ""
		self.keywordsEn = ""
		self.pages = ""
		self.fileLabel = ""
		self.fileLink = ""
		self.autores = []
		self.referencias = []

		titulo = ""
		autores = ""
		logging.info("Artigo")
		if not pdf_text_ is None:
			#logging.debug("debug:'%s'"%pdf_text_)
			achou = False
			stri = StringIO(pdf_text_)

			# busca pelo título (presume-se que a primeira linha é o título)
			linha = stri.readline().strip()
			while linha == "" or "SIBGRAPI" in linha:
				# alguns pdfs abrem com a informação "Instruções aos Autores de Contribuições para o SIBGRAPI "
				linha = stri.readline().strip()
				logging.info("\tlinha1: {}".format(linha))

			# for i in range(10):
			# 	linha = stri.readline()
			# 	print("linha {}: {}".format(i, linha))
			#sys.exit()
			titulo = linha
			# incorpora as próximas linhas não nulas para os casos de títulos em múltiplas linhas
			while linha != "" and not linha[-1].isnumeric() and not linha[-2].isnumeric():
				linha = stri.readline().strip()
				logging.info("\tlinha2.1: {}".format(linha))
				if len(linha) > 0:
					if not linha[-1].isnumeric() and not linha[-2].isnumeric():
						titulo += " " + linha
					else:
						break

				linha = stri.readline().strip()
				logging.info("\tlinha2.2: {}".format(linha))
				if len(linha) > 0:
					if not linha[-1].isnumeric() and not linha[-2].isnumeric():
						titulo += linha
					else:
						break

			logging.info("Título: {}".format(titulo))

			# busca linha dos autores
			while linha == "":
				linha = stri.readline().strip()
				logging.info("\tlinha3: {}".format(linha))

			autores = linha
			# caso de autores em múltiplas linhas
			while linha != "" and not linha[1].isnumeric() :
				linha = stri.readline().strip()
				logging.info("\tlinha4.1: {}".format(linha))
				if not "Instituto" in linha and not "Univ" in linha and not " RS"  in linha and not " SC"  in linha:
					autores += " " + linha

				linha = stri.readline().strip()
				logging.info("\tlinha4.2: {}".format(linha))

			logging.info("Autores antes processamento: {}".format(autores))
			# remove supperscript para instituição (superestimado até 30 instituições)
			for i in range(30):
				autores = autores.replace("%d"%i,"")
				autores = autores.replace(",,", ",")
				# alguns autores usam 'e' como separador ao invés de ','
				autores = autores.replace(" e ", ", ")
				autores = autores.replace(" and ", ", ")
			autores = autores.strip()
			# remove última virgula, caso necessário
			if autores[-1] == ",":
				autores = autores[:-1]
			logging.info("Autores após processamento: {}".format(autores))

			#sys.exit()
			while True:
				linha = stri.readline()
				logging.debug(linha)
				if linha == "Referências\n" or linha == "Referências Bibliográficas\n":
					logging.info("achou linha Referências!")
					achou = True
					break

				if not linha:
					break

			if not achou:
				logging.info("não achou linha Referências!")
				stri = StringIO(pdf_text_)
				while True:
					linha = stri.readline()
					logging.debug(linha)
					if linha == "Referências\n" or "Referências" or "References" in linha:
						logging.info("achou linha Referências!")
						achou = True
						break

					if not linha:
						break

			if not achou:
				logging.error("ainda não achou linha Referências!")
				sys.exit(-1)

			else:
				linha = stri.readline().strip()
				logging.debug("linha_lida:'%s'" % linha)
				referencia = ""
				cont_ref = 0  # contador de referências
				cont_null = 0  # contador de linhas nulas para rede de segurança; após 25 linhas nulas o algoritmo para.
				primeiro = True  # variável de apoio para decidir sobre casos específico de referências mal formatadas

				while True:
					if "Introdução" in linha:
						pass

					if linha == "":
						cont_null += 1
						logging.debug("análise: linha nula -> pulando e contando... cont_null:%d" % cont_null)
						linha = stri.readline().strip()
						logging.debug("linha_lida:'%s'" % linha)

					else:
						cont_null = 0  # linha não nula reinicia contador de linhas nulas

						# laço para montar uma referência completa
						while True:

							if linha == "":
								logging.debug("análise: linha nula -> pulando e contando... cont_null:%d" % cont_null)
								cont_null += 1

								linha = stri.readline().strip()
								logging.debug("linha_lida:'%s'" % linha)

								# nova linha em branco e a referência parece ter chegado ao fim
								if referencia[-1] == '.':

									if linha[0:2].upper() == "IN":
										# trata-se de uma continuação, por exemplo:
										# FERRARA, E.; Varol, O.; Davis, C; Menczer, F; Flammini, A. The rise of social bots.
										# In: Communications of the ACM, v59, n.2, 2016.
										pass

									elif linha[0:7].upper() == "REVISTA":
										# trata-se de uma continuação, por exemplo:
										# SASTRE , Angelo; CORREIO , Claudia Silene Pereira de Oliveira;  CORREIO , Francisco Rolfsen Belda. A  influência do “filtro bolha” na difusão de  Fake News nas mídias sociais: reflexões sobre as mudanças nos  algoritmos do Facebook.
										# Revista GEMInIS, São Carlos, UFSCar, v. 9, n. 1, pp.4-17, jan. / abr. 2018.
										pass

									elif linha[0:4].upper() == "HTTP":
										# trata-se de uma continuação, por exemplo:
										# CGEE (2012). Redes elétricas inteligentes: contexto naional.
										# http://www.cgee.org.br/atividades/redirect/8050. Acessado em Outubro/2017.
										pass

									elif linha[0:3].upper() == "ED.":
										# Kurose, J. F. (2013). Redes de Computadores e a Internet: uma abordagem topdown. 6.
										# ed. São Paulo: Editora Pearson.
										pass

									elif linha[0:11].upper() == "ACESSADO EM":
										# trata-se de uma continuação, por exemplo:
										# ....
										# Acessado em Outubro/2017.
										pass

									elif linha[0:9].upper() == "ACESSO EM":
										# trata-se de uma continuação, por exemplo:
										# Sophos XG. (2017). Disponível em: https://www.m3corp.com.br/sophos/sophos-utm-2.
										# Acesso em: 26/06/2019 Sophos. (2017). Disponível em: <https://www.sophos.com/en-us.aspx. Acesso em: 26/06/2019 Turban, E.; Volonino, L. (2013). Tecnologia da Informação para Gestão: Em busca do melhor desempenho Estratégico e Operacional. 8ed. Porto Alegre: Bookman. 721 p.
										pass

									elif linha[0:11].upper() == "DISSERTACAO" or linha[0:11].upper() == "DISSERTAÇÃO":
										# trata-se de uma continuação, por exemplo:
										# Panes, G. G. (2011). Firewall Dinâmico: Uma implementação Cliente/Servidor.
										# Dissertação de Mestrado, Pós-Graduação em Ciência da Computação, 72p.
										pass

									elif linha[0:10].upper() == "DISPONIVEL":
										# trata-se de uma continuação, por exemplo:
										# Bär, H. (2017). 4 vulnerabilidades que mais afetam a segurança da informação.
										# Disponivel em: https://triplait.com/4-vulnerabilidades-que-mais-afetam-a-segurancada-informacao. Acesso em:  25/05/2019.
										pass

									else:
										break

								if cont_null > 25:
									break

							# HACK para um caso particular
							# TODO: generalizar solução
							elif (len(linha) > 9 and "Turban, E." == linha[0:10]):
								logging.debug("\t\t análise: referência nova após referência sem ponto.")
								referencia = "%s." % referencia
								linha = " %s" % linha
								#linha = stri.readline().strip()
								#logging.debug("linha_lida:'%s'" % linha)
								break

							# casos particulares; pode acontecer de duas duas referências não terem linhas em branco intermediárias
							# a menos que a referência atual termine em ',' ou ':' e então a próxima linha é uma continuação
							elif linha[0].isupper() and \
									(".," in linha or " and " in linha ) and \
									primeiro == False \
									and not "pages" in linha and \
									not "Proceedings" in referencia and \
									not linha[0:4].upper() == "HTTP" and \
									not linha[0:10].upper() == "DISPONIVEL" and \
									not linha[0:10].upper() == "DISPONÍVEL" and \
									not linha[0:11].upper() == "DISSERTACAO" and \
									not linha[0:11].upper() == "DISSERTAÇÃO" and \
									not linha[0:11].upper() == "ACESSADO EM" and \
									not linha[0:9].upper() == "ACESSO EM" and \
									(len(referencia) > 1 and referencia[-1] != ',') and \
								    (len(referencia) > 1 and referencia[-1] != ':') and \
								    (len(referencia) > 1 and referencia[-1] == '.') :
								logging.debug("\t\t análise: linha upper e com '.,'")

								break

							else:

								if len(referencia) > 1 and referencia[-1] == '-':
									logging.debug("\t\t análise: linha com dados terminado com -" )
									referencia = "%s%s" % (referencia[0:-1], linha.strip())

								else:
									logging.debug("\t\t análise: linha com dados normal")
									referencia = "%s %s" % (referencia, linha.strip())

								primeiro = False
								logging.debug("\t\t referência_temporária:'%s" % referencia)

								linha = stri.readline().strip()
								logging.debug("linha_lida:'%s'" % linha)

						cont_ref += 1

						logging.info("\n cont_ref:%d referência_final:'%s'\n" %(cont_ref, referencia))

						# teste de segurança																	''
						if referencia != "":
							if "1. Introdução" in referencia and "Referência" in referencia:
								# evitar alguns casos especiais onde a estrutura do artigo aparece após referências
								pass
							else:
								referencia = referencia.strip()
								self.referencias.append(Referencia(self.seq, referencia))

						# reinicia variáveis de controle
						referencia = ""
						primeiro = True

					if cont_null > 25:
						break

		#print(pdf_text_)
		#sys.exit()

		self.fileLink = os.path.basename(pdf_file_name_)
		if not bib_database_ is None:
			for attribute in self.__dict__.keys():
				if attribute in bib_database_.entries[0]:
					value = bib_database_.entries[0][attribute]
					self.__setattr__(attribute, value)
					logging.debug("Attribute: %s [OK] Value: %s" % (attribute, value))
				else:
					logging.debug("Attribute: %s [MISS]" % attribute)

			#link não é necessário, será gerado pelo sistema OpenLib
			#self.fileLink = bib_database_.entries[0]["url"]
			conta_autor = 0
			autores_lista = bib_database_.entries[0]["author"].split(" and ")
			logging.debug(autores_lista)
			for autor_str in autores_lista:
				conta_autor += 1
				autor_str = autor_str.strip()
				logging.info("\t\t(%d/%d) %s" % (conta_autor, len(autores_lista), autor_str))
				autor = Autor(self.seq, autor_str)
				logging.debug(autor)
				self.autores.append(autor)
		else:
			self.title = titulo
			conta_autor = 0
			autores_lista = autores.split(",")
			logging.info(autores_lista)
			for autor_str in autores_lista:
				conta_autor += 1
				autor_str = autor_str.strip()
				logging.info("\t\t(%d/%d) %s" % (conta_autor, len(autores_lista), autor_str))
				autor = Autor(self.seq, autor_str)
				logging.debug(autor)
				self.autores.append(autor)




	def __str__(self):
		msg = "\n"
		msg += "\tArtigo\n"
		msg += "\t-------\n"
		msg += "\t\tseq: %d\n" % self.seq
		msg += "\t\tlanguage:%s\n" % self.language
		msg += "\t\tsectionAbbrev:%s\n" % self.sectionAbbrev
		msg += "\t\ttitle:%s\n" % self.title
		msg += "\t\ttitleEn:%s\n" % self.titleEn
		msg += "\t\tkeywords:%s\n" % self.keywords
		msg += "\t\tkeywordsEn:%s\n" % self.keywordsEn
		msg += "\t\tfileLink:%s\n" % self.fileLink
		#msg += "\tautores:%s\n" % self.autores
		for autor in self.autores:
			msg += autor.__str__()
		msg += '\n'
		for referencia in self.referencias:
			msg += referencia.__str__()
		msg += '\n'
		return msg


def exporta_artigos_xlsx(planilha_, artigos_, acrescentar_=True):
	'''
	Preenche planilha de artigos com dados

	:param planilha_: openpyxl.sheet
	:param dados_detalhe_grupos_: dicionário com valores
	'''

	linha = gera_cabecalho(planilha_, CAMPOS_ARTIGOS, 1)
	coluna = 1

	for artigo in artigos_:
		coluna = 1
		for campo in CAMPOS_ARTIGOS:
			valor = artigo.__getattribute__(campo)
			planilha_.cell(row=linha, column=coluna).value = valor
			logging.debug("artigo: %d campo: %s valor: %s" % (artigo.seq, campo, valor))
			coluna += 1
		linha += 1


def exporta_autores_xlsx(planilha_, artigos_, acrescentar_=True):
	'''
	Preenche planilha de autores com dados

	:param planilha_: openpyxl.sheet
	:param dados_detalhe_grupos_: dicionário com valores
	'''

	linha = gera_cabecalho(planilha_, CAMPOS_AUTORES, 1)
	coluna = 1

	for artigo in artigos_:
		for autor in artigo.autores:
			coluna = 1

			for campo in CAMPOS_AUTORES:
				valor = autor.__getattribute__(campo)
				planilha_.cell(row=linha, column=coluna).value = valor
				logging.debug("artigo: %d campo: %s valor: %s" % (artigo.seq, campo, valor))
				coluna += 1
			linha += 1


def exporta_referencias_xlsx(planilha_, artigos_, acrescentar_=True):
	'''
	Preenche planilha de autores com dados

	:param planilha_: openpyxl.sheet
	:param dados_detalhe_grupos_: dicionário com valores
	'''

	linha = gera_cabecalho(planilha_, CAMPOS_REFERENCIAS, 1)
	coluna = 1

	for artigo in artigos_:
		for referencia in artigo.referencias:
			coluna = 1

			for campo in CAMPOS_REFERENCIAS:
				valor = referencia.__getattribute__(campo)
				planilha_.cell(row=linha, column=coluna).value = valor
				logging.debug("artigo: %d campo: %s valor: %s" % (artigo.seq, campo, valor))
				coluna += 1
			linha += 1

def exporta_secoes_xlsx(planilha_, secao_, acrescentar_=True):
	'''
	Exporta dados de seção para planilha
	:param planilha_:
	:param secao_:
	:param acrescentar_:
	:return:
	'''

	linha = gera_cabecalho(planilha_, CAMPOS_SECOES, 3)
	coluna = 1

	for campo in CAMPOS_SECOES:
		valor = secao_.__getattribute__(campo)
		planilha_.cell(row=linha, column=coluna).value = valor
		logging.debug("campo: %s valor: %s" % (campo, valor))
		coluna += 1


def gera_cabecalho(planilha_, campos_, coluna_indice_):
	linha = 1
	coluna = 1

	if planilha_.cell(row=linha, column=coluna).value is None:
		for campo in campos_:
			planilha_.cell(row=linha, column=coluna).value = campo
			coluna += 1
		linha += 1

	coluna = coluna_indice_
	while not planilha_.cell(row=linha, column=coluna).value is None:
		linha += 1

	return linha


def gera_workbook_planilha(nome_arquivo_, acrescentar_=True):

	logging.debug("def_gera_workbook_planilha  nome_arquivo_:%s acrescentar_:%s"%(nome_arquivo_, acrescentar_))
	if Path(nome_arquivo_).is_file() and not acrescentar_:
		logging.debug("removendo arquivo...")
		os.remove(nome_arquivo_)
		logging.debug("pronto.")

	if not Path(nome_arquivo_).is_file():
		logging.debug("arquivo não existe! criando novo arquivo...")
		workbook = Workbook()
		workbook.remove(workbook.active)
		workbook.create_sheet(PLANILHA_PADRAO)
		workbook.save(nome_arquivo_)
		workbook.close()

	workbook = load_workbook(nome_arquivo_)
	planilha = workbook[PLANILHA_PADRAO]

	return workbook, planilha


def le_seq_artigo(nome_arquivo_, acrescentar_=True):
	logging.debug("le_seq_artigo  nome_arquivo_:%s acrescentar_:%s" % (nome_arquivo_, acrescentar_))
	if not acrescentar_ or not Path(nome_arquivo_).is_file():
		seq_artigo = 1
	else:
		workbook = load_workbook(nome_arquivo_)
		planilha = workbook[PLANILHA_PADRAO]
		coluna = 1
		linha = 1

		while not planilha.cell(row=linha, column=coluna).value is None:
			logging.debug("Linha: %d Valor:%s" % (linha, planilha.cell(row=linha, column=coluna).value))
			linha += 1

		seq_artigo = linha - 1  # descontar o cabeçalho

	return seq_artigo


def main():
	'''
	Programa principal
	'''

	# Configura argumentos
	parser = argparse.ArgumentParser(description='Gera arquivos de produção bibliográfica de eventos da SBC.')
	parser.add_argument("--dir", "-d", help="diretório com arquivos de entrada.", default=DIRETORIO_ENTRADA_PADRAO)
	parser.add_argument("--secao", "-s", help="abreviatura da seção a ser processada (coluna sectionAbbrev do arquivo Secoes.xlsx), padrão: diretório de entrada.", default=None)
	parser.add_argument('--acrescentar', dest='acrescentar',  help="acrescentar aos arquivos pré-existentes.", action='store_true', default=True)
	parser.add_argument('--nao-acrescentar', dest='acrescentar', help="sobreescrever arquivos pré-existentes.", action='store_false')

	parser.add_argument("--artigos", "-a", help="planilha de saída com artigos.", default=PLANILHA_ARTIGOS_PADRAO)
	parser.add_argument("--autores", "-u", help="planilha de saída com autores.", default=PLANILHA_AUTORES_PADRAO)
	parser.add_argument("--referencias", "-r", help="planilha de saída com referências.", default=PLANILHA_REFERENCIAS_PADRAO)
	parser.add_argument("--secoes", "-e", help="planilha de saída com seções.", default=PLANILHA_SECOES_PADRAO)

	help_log = "nível de log (INFO=%d DEBUG=%d)"%(logging.INFO, logging.DEBUG)
	parser.add_argument("--log", "-l", help=help_log, default=logging.INFO, type=int)
	parser.print_help()

	# lê argumentos da linha de comando
	args = parser.parse_args()
	if args.secao is None:
		args.secao = args.dir
		
	# configura log
	#logging.basicConfig(level=args.log, format='%(asctime)s - %(message)s')
	logging.basicConfig(level=args.log, format='%(message)s')

	# mostra parâmetros de entrada
	logging.info("")
	logging.info("PARÂMETROS DE ENTRADA")
	logging.info("---------------------")

	logging.info("\tdir        : %s" % args.dir)
	logging.info("\tseção      : %s" % args.secao)
	logging.info("\tartigos    : %s" % args.artigos)
	logging.info("\tautores    : %s" % args.autores)
	logging.info("\treferencias: %s" % args.referencias)
	logging.info("\tsecoes     : %s" % args.secoes)

	logging.info("\tlog        : %d" % args.log)
	logging.info("\tacrescentar: %s" % args.acrescentar)

	# mostra parâmetros calculados
	logging.info("")
	logging.info("PARÂMETROS CALCULADOS")
	logging.info("---------------------")

	arquivos_pdfs = [arquivo for arquivo in Path(args.dir).rglob('[!~]*.pdf')]
	logging.info("\tpdfs       : %s" % arquivos_pdfs)

	# inicializa variáveris
	artigos = []
	conta_arquivo = 0

	# processa dados
	logging.info("")
	logging.info("LEITURA DE DADOS")
	logging.info("----------------")

	try:
		sequencia = le_seq_artigo(args.artigos, args.acrescentar)
		logging.debug("Sequencia: %d" %sequencia)
		for posix_path_pdf in arquivos_pdfs:
			conta_arquivo += 1

			nome_arquivo_pdf = str(posix_path_pdf)
			nome_arquivo_bib = nome_arquivo_pdf.replace(".pdf", ".bib")

			bib_database = None
			if not os.path.isfile(nome_arquivo_bib):
				logging.exception("\t(%d/%d) PDF:%s [OK] BIB:%s [NOT FOUND] " % (conta_arquivo, len(arquivos_pdfs), nome_arquivo_pdf, nome_arquivo_bib))

			else:
				logging.info("\t(%d/%d) PDF:%s [OK] BIB:%s [OK]"%(conta_arquivo, len(arquivos_pdfs), nome_arquivo_pdf, nome_arquivo_bib))

				bib_text_parser = BibTexParser()
				bib_text_parser.customization = convert_to_unicode
				bibtex_file = open(nome_arquivo_bib)
				bib_database = bibtexparser.load(bibtex_file, parser=bib_text_parser)

			try:
				#pdf_texto = pdf_to_text(nome_arquivo_pdf)
				#print(pdf_texto)

				pdf_dados = tika_parser.from_file(nome_arquivo_pdf)

				pdf_texto = pdf_dados['content']
				#print("Title: {}".format(pdf_dados['tile']))
				#print("Authors: {}".format(pdf_dados['author']))
				#pdf_texto = pdf_texto.encode('utf-8', errors='ignore')

				artigo = Artigo(LANGUAGE_PADRAO, args.secao, sequencia, bib_database, pdf_texto, nome_arquivo_pdf)
				sequencia += 1
				logging.info(artigo)
				artigos.append(artigo)

			except Exception as e:
				print(e)
				sys.exit(-1)
				#artigo = Artigo(LANGUAGE_PADRAO, args.secao, sequencia)
				#artigos.append(artigo)

	except Exception as e:
		print(e)
		#sys.exit(-1)


	logging.info("")
	logging.info("RESULTADOS")
	logging.info("----------")

	logging.info("")
	logging.info("\tSEÇÃO")
	logging.info("\t-----")
	logging.info("\t\tprocessando...")
	workbook, planilha = gera_workbook_planilha(args.secoes, args.acrescentar)
	secao = Secao(args.secao)
	exporta_secoes_xlsx(planilha, secao)
	logging.info("\t\tgravando...")
	workbook.save(args.secoes)
	logging.info("\t\tpronto.")

	logging.info("")
	logging.info("\tARTIGOS")
	logging.info("\t----------")
	logging.info("\t\tprocessando...")
	workbook, planilha = gera_workbook_planilha(args.artigos, args.acrescentar)
	exporta_artigos_xlsx(planilha, artigos)
	logging.info("\t\tgravando...")
	workbook.save(args.artigos)
	logging.info("\t\tpronto.")

	logging.info("")
	logging.info("\tAUTORES")
	logging.info("\t----------")
	logging.info("\t\tprocessando...")
	workbook, planilha = gera_workbook_planilha(args.autores, args.acrescentar)
	exporta_autores_xlsx(planilha, artigos)
	logging.info("\t\tgravando...")
	workbook.save(args.autores)
	logging.info("\t\tpronto.")

	logging.info("")
	logging.info("\tREFERÊNCIAS")
	logging.info("\t-----------")
	logging.info("\t\tprocessando...")
	workbook, planilha = gera_workbook_planilha(args.referencias, args.acrescentar)
	exporta_referencias_xlsx(planilha, artigos)
	logging.info("\t\tgravando...")
	workbook.save(args.referencias)
	logging.info("\t\tpronto.")




	logging.info("\nfim.\n")

if __name__ == "__main__":
	main()
